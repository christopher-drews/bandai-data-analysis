"""Extract USD exchange rates by month from the Bandai royalty workbook.

Walks every monthly sheet in BNEPA_Royalty_Report_*.xlsx, reads the
"Sales Period" dates in B3/B4 to determine the month(s) the sheet covers,
and locates the exchange rate (cell position varies across template
revisions). Writes one row per calendar month to
data/level_0_extract_exchange_rates/exchange_rates.csv.
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

DEFAULT_WORKBOOK = Path("BNEPA_Royalty_Report_MAY2026_LV.xlsx")
DEFAULT_OUTPUT = Path("data/level_0_extract_exchange_rates/exchange_rates.csv")

# Sheets that don't carry monthly royalty data.
NON_MONTHLY_SHEETS = {"New template", "銷售庫存統計總表"}

# Cells to probe for the exchange-rate value, in order. The label sits in a
# neighboring cell that differs by template revision; the value itself is
# always one of these.
RATE_CELL_CANDIDATES = ("C6", "B7", "D7")

# (start_cell, end_cell) pairs for the sales period, in order. Newer template
# revisions put the dates in column A; older ones in column B.
PERIOD_CELL_CANDIDATES = (("A4", "A5"), ("B3", "B4"))


def find_period(ws) -> tuple[datetime, datetime] | None:
    for start_coord, end_coord in PERIOD_CELL_CANDIDATES:
        start = ws[start_coord].value
        end = ws[end_coord].value
        if isinstance(start, datetime) and isinstance(end, datetime):
            return start, end
    return None


def month_range(start: datetime, end: datetime) -> list[str]:
    """Return YYYY-MM strings for every month between start and end inclusive."""
    months: list[str] = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def extract_rate(ws) -> float | None:
    for coord in RATE_CELL_CANDIDATES:
        value = ws[coord].value
        if isinstance(value, (int, float)):
            return float(value)
    return None


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, type=Path)
    parser.add_argument("--output", default=DEFAULT_OUTPUT, type=Path)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.workbook, read_only=False, data_only=True)

    rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in NON_MONTHLY_SHEETS:
            continue
        ws = wb[sheet_name]

        period = find_period(ws)
        if period is None:
            print(f"  skip {sheet_name!r}: no sales period found", file=sys.stderr)
            continue
        start, end = period

        rate = extract_rate(ws)
        if rate is None:
            print(f"  skip {sheet_name!r}: no rate at {RATE_CELL_CANDIDATES}", file=sys.stderr)
            continue

        for month in month_range(start, end):
            rows.append({
                "month": month,
                "exchange_rate": rate,
                "sheet_name": sheet_name.strip(),
            })

    rows.sort(key=lambda r: r["month"])

    args.output.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["month", "exchange_rate", "sheet_name"]
    with args.output.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {args.output} ({len(rows)} rows)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
