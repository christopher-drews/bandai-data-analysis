"""Extract per-month SRPs from the Bandai royalty workbook and join PAX codes.

Walks every monthly sheet in BNEPA_Royalty_Report_*.xlsx, pulls the
``SRP (CNY)`` column for each product, normalizes the product names via
``normalize.normalize_name``, joins to ``royalty_pax_match.xlsx`` for the
``paxCode`` and ``Customer Reference`` columns, and collapses consecutive
same-SRP observations into start/end month ranges.

Output: product_srp_history.csv with columns
    Product Name, Normalized Name, SRP, currency, start_month, end_month,
    paxCode, Customer Reference
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

from normalize import normalize_name

DEFAULT_WORKBOOK = Path("BNEPA_Royalty_Report_MAY2026_LV.xlsx")
DEFAULT_PAX_MATCH = Path("royalty_pax_match.xlsx")
DEFAULT_OUTPUT = Path("product_srp_history.csv")

SKIP_SHEETS = {"New template", "銷售庫存統計總表"}
HEADER_SCAN_DEPTH = 12
PERIOD_CELL_CANDIDATES = (("A4", "A5"), ("B3", "B4"))
SRP_CURRENCY = "CNY"


def find_period(ws) -> tuple[datetime, datetime] | None:
    for start_coord, end_coord in PERIOD_CELL_CANDIDATES:
        start = ws[start_coord].value
        end = ws[end_coord].value
        if isinstance(start, datetime) and isinstance(end, datetime):
            return start, end
    return None


def month_range(start: datetime, end: datetime) -> list[str]:
    months: list[str] = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def find_header_row(path: Path, sheet: str, depth: int = HEADER_SCAN_DEPTH) -> int:
    probe = pd.read_excel(path, sheet_name=sheet, header=None, nrows=depth, engine="openpyxl")
    for i in range(len(probe)):
        cells = {str(v).strip() for v in probe.iloc[i].tolist() if pd.notna(v)}
        if "Product Name" in cells:
            return i
    raise ValueError(f"No 'Product Name' header found in first {depth} rows of {sheet!r}")


def build_pax_lookup(path: Path) -> dict[str, tuple[str, str]]:
    """normalize_name(Name) -> (paxCode, customer_reference). Last-write wins on dupes."""
    df = pd.read_excel(path, sheet_name="SKUs", engine="openpyxl")
    df = df[["Name", "PAX", "Customer Reference", "fallback_pax_code"]].copy()
    df["resolved_pax"] = df["PAX"].fillna(df["fallback_pax_code"])
    df["slug"] = df["Name"].map(normalize_name)
    df = df[df["slug"] != ""]

    lookup: dict[str, tuple[str, str]] = {}
    for slug, group in df.groupby("slug"):
        if len(group) > 1:
            print(f"  warn: duplicate slug {slug!r} in pax-match ({len(group)} rows)", file=sys.stderr)
        row = group.iloc[0]
        pax = row["resolved_pax"] if pd.notna(row["resolved_pax"]) else ""
        ref = row["Customer Reference"] if pd.notna(row["Customer Reference"]) else ""
        lookup[slug] = (pax, ref)
    return lookup


def extract_sheet_srps(path: Path, sheet: str) -> pd.DataFrame:
    """Per-row Product Name + SRP (CNY) from one sheet, normalized and cleaned."""
    hdr = find_header_row(path, sheet)
    df = pd.read_excel(path, sheet_name=sheet, header=hdr, engine="openpyxl")
    cols = {c.strip(): c for c in df.columns if isinstance(c, str)}
    pn = cols.get("Product Name")
    srp = cols.get("SRP (CNY)")
    if not (pn and srp):
        return pd.DataFrame(columns=["Product Name", "Normalized Name", "SRP"])

    sub = df[[pn, srp]].rename(columns={pn: "Product Name", srp: "SRP"})
    sub["Product Name"] = sub["Product Name"].astype(str).str.strip()
    sub = sub[sub["Product Name"].ne("") & sub["Product Name"].ne("nan")]
    sub["SRP"] = pd.to_numeric(sub["SRP"], errors="coerce")
    sub = sub.dropna(subset=["SRP"])
    sub["Normalized Name"] = sub["Product Name"].map(normalize_name)
    sub = sub[sub["Normalized Name"] != ""]
    return sub[["Product Name", "Normalized Name", "SRP"]]


def collapse_runs(per_month: pd.DataFrame, sheet_order: list[str]) -> pd.DataFrame:
    """Collapse consecutive same-SRP sheets into (start_month, end_month) runs.

    Non-consecutive sheets (a gap in coverage) start a new run even if the SRP
    is unchanged — matches the existing notebook behavior.
    """
    sheet_idx = {s: i for i, s in enumerate(sheet_order)}
    per_month = per_month.copy()
    per_month["_idx"] = per_month["sheet"].map(sheet_idx)
    per_month = per_month.sort_values(["Normalized Name", "_idx"]).reset_index(drop=True)

    runs: list[dict] = []
    for slug, grp in per_month.groupby("Normalized Name", sort=False):
        grp = grp.sort_values("_idx")
        run_srp = None
        run_start_month = run_end_month = None
        prev_idx = None
        product_name = None
        for _, row in grp.iterrows():
            if run_srp is None:
                run_srp = row["SRP"]
                run_start_month = row["start_month"]
                run_end_month = row["end_month"]
                prev_idx = row["_idx"]
                product_name = row["Product Name"]
                continue
            if row["SRP"] == run_srp and row["_idx"] == prev_idx + 1:
                run_end_month = row["end_month"]
            else:
                runs.append({
                    "Product Name": product_name,
                    "Normalized Name": slug,
                    "SRP": run_srp,
                    "start_month": run_start_month,
                    "end_month": run_end_month,
                })
                run_srp = row["SRP"]
                run_start_month = row["start_month"]
                run_end_month = row["end_month"]
                product_name = row["Product Name"]
            prev_idx = row["_idx"]
        if run_srp is not None:
            runs.append({
                "Product Name": product_name,
                "Normalized Name": slug,
                "SRP": run_srp,
                "start_month": run_start_month,
                "end_month": run_end_month,
            })
    return pd.DataFrame(runs)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, type=Path)
    parser.add_argument("--pax-match", default=DEFAULT_PAX_MATCH, type=Path)
    parser.add_argument("--output", default=DEFAULT_OUTPUT, type=Path)
    args = parser.parse_args()

    pax_lookup = build_pax_lookup(args.pax_match)
    print(f"Loaded {len(pax_lookup)} PAX-match entries", file=sys.stderr)

    wb = openpyxl.load_workbook(args.workbook, read_only=False, data_only=True)
    monthly = [s for s in wb.sheetnames if s.strip() not in SKIP_SHEETS]

    sheet_periods: dict[str, tuple[datetime, datetime]] = {}
    for sheet in monthly:
        period = find_period(wb[sheet])
        if period is None:
            print(f"  skip {sheet!r}: no sales period found", file=sys.stderr)
            continue
        sheet_periods[sheet] = period

    sheet_order = sorted(sheet_periods, key=lambda s: sheet_periods[s][0])

    parts: list[pd.DataFrame] = []
    for sheet in sheet_order:
        start, end = sheet_periods[sheet]
        months = month_range(start, end)
        sub = extract_sheet_srps(args.workbook, sheet)
        if sub.empty:
            print(f"  skip {sheet!r}: no SRP rows", file=sys.stderr)
            continue
        # If the same normalized product appears multiple times in one sheet,
        # keep the modal SRP and the first display name (matches notebook).
        sub = (
            sub.groupby("Normalized Name", as_index=False)
            .agg(**{
                "Product Name": ("Product Name", "first"),
                "SRP": ("SRP", lambda s: s.mode().iloc[0]),
            })
        )
        sub["sheet"] = sheet
        sub["start_month"] = months[0]
        sub["end_month"] = months[-1]
        parts.append(sub)

    per_month = pd.concat(parts, ignore_index=True)
    runs = collapse_runs(per_month, sheet_order)

    # One canonical display name per slug — the spelling from the most recent
    # sheet that carried it. Avoids stray casing/typos in older sheets leaking
    # into the Product Name column.
    sheet_rank = {s: i for i, s in enumerate(sheet_order)}
    display_names = (
        per_month.assign(_rank=per_month["sheet"].map(sheet_rank))
        .sort_values("_rank")
        .drop_duplicates(subset=["Normalized Name"], keep="last")
        .set_index("Normalized Name")["Product Name"]
    )
    runs["Product Name"] = runs["Normalized Name"].map(display_names)

    runs["currency"] = SRP_CURRENCY
    runs["paxCode"] = ""
    runs["Customer Reference"] = ""
    missing: list[str] = []
    for i, slug in enumerate(runs["Normalized Name"]):
        entry = pax_lookup.get(slug)
        if entry is None:
            missing.append(runs.at[i, "Product Name"])
            continue
        runs.at[i, "paxCode"] = entry[0]
        runs.at[i, "Customer Reference"] = entry[1]

    runs = runs[[
        "Product Name", "Normalized Name", "SRP", "currency",
        "start_month", "end_month", "paxCode", "Customer Reference",
    ]].sort_values(["Product Name", "start_month"]).reset_index(drop=True)

    runs.to_csv(args.output, index=False)
    print(f"Wrote {args.output} ({len(runs)} rows)", file=sys.stderr)

    if missing:
        unique_missing = sorted(set(missing))
        print(f"\n{len(unique_missing)} product(s) had no PAX-match entry:", file=sys.stderr)
        for name in unique_missing:
            print(f"  - {name}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
