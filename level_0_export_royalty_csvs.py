"""Export each monthly period sheet of the Bandai royalty workbook as its own CSV.

Walks every period sheet in BNEPA_Royalty_Report_*.xlsx, dumps the data table
(from the ``Product Name`` header row down, all columns preserved) into
``data/level_0_export_royalty_csvs/<YYYY-MM>.csv``. Sheets covering multiple
months produce ``<YYYY-MM>_<YYYY-MM>.csv``.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

DEFAULT_WORKBOOK = Path("BNEPA_Royalty_Report_MAY2026_LV.xlsx")
DEFAULT_OUTPUT_DIR = Path("data/level_0_export_royalty_csvs")

SKIP_SHEETS = {"New template", "銷售庫存統計總表"}
HEADER_SCAN_DEPTH = 12
PERIOD_CELL_CANDIDATES = (("A4", "A5"), ("B3", "B4"))


def find_period(ws) -> tuple[datetime, datetime] | None:
    for start_coord, end_coord in PERIOD_CELL_CANDIDATES:
        start = ws[start_coord].value
        end = ws[end_coord].value
        if isinstance(start, datetime) and isinstance(end, datetime):
            return start, end
    return None


def find_header_row(path: Path, sheet: str, depth: int = HEADER_SCAN_DEPTH) -> int:
    probe = pd.read_excel(path, sheet_name=sheet, header=None, nrows=depth, engine="openpyxl")
    for i in range(len(probe)):
        cells = {str(v).strip() for v in probe.iloc[i].tolist() if pd.notna(v)}
        if "Product Name" in cells:
            return i
    raise ValueError(f"No 'Product Name' header found in first {depth} rows of {sheet!r}")


def period_filename(start: datetime, end: datetime) -> str:
    if start.year == end.year and start.month == end.month:
        return f"{start:%Y-%m}.csv"
    return f"{start:%Y-%m}_{end:%Y-%m}.csv"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, type=Path)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, type=Path)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(args.workbook, read_only=False, data_only=True)
    sheets = [s for s in wb.sheetnames if s.strip() not in SKIP_SHEETS]

    written = 0
    for sheet in sheets:
        period = find_period(wb[sheet])
        if period is None:
            print(f"  skip {sheet!r}: no sales period found", file=sys.stderr)
            continue
        start, end = period

        try:
            hdr = find_header_row(args.workbook, sheet)
        except ValueError as e:
            print(f"  skip {sheet!r}: {e}", file=sys.stderr)
            continue

        df = pd.read_excel(args.workbook, sheet_name=sheet, header=hdr, engine="openpyxl")
        df = df.dropna(how="all")

        out_path = args.output_dir / period_filename(start, end)
        df.to_csv(out_path, index=False)
        print(f"Wrote {out_path} ({len(df)} rows)", file=sys.stderr)
        written += 1

    print(f"\nWrote {written} CSV(s) to {args.output_dir}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
